from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.conf import settings
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.exceptions import ValidationError
from django.db.models import Q, Sum
import csv
from .models import BorrowRecord, BookRequest
from ..books.models import Book
from ..users.models import User
from ..users.permissions import (
    admin_required, librarian_or_admin_required, student_required,
    check_librarian_or_admin, check_student, check_admin
)


@login_required
def borrow_book(request, book_pk):
    return redirect('books:list')


@login_required
@librarian_or_admin_required
def return_book(request, record_pk):
    """Librarian/Admin processes book return"""
    record = get_object_or_404(BorrowRecord, pk=record_pk)
    
    if record.status == 'returned':
        messages.warning(request, 'This book has already been returned.')
        return redirect('borrow:issue_return')
    
    # Process return
    record.return_book()
    
    # Update returned_to field using update to bypass validation
    BorrowRecord.objects.filter(pk=record.pk).update(returned_to=request.user)
    
    # Refresh from database
    record.refresh_from_db()
    
    messages.success(request, f'Book "{record.book.title}" has been returned by {record.user.username}.')

    # Show fine information if applicable
    if record.fine_amount > 0:
        messages.warning(request, f'Fine amount: ETB {record.fine_amount}. Please collect the fine.')

    # Show waitlist notification info
    from apps.borrow.models import BookRequest
    waitlist_count = BookRequest.objects.filter(
        book=record.book, status='pending'
    ).count()
    if waitlist_count > 0:
        messages.info(request, f'Waitlist notification sent to {waitlist_count} user(s) waiting for "{record.book.title}".')

    return redirect('borrow:issue_return')


@login_required
@student_required
def student_return_book(request, record_pk):
    """Student returns their own borrowed book (digital library)"""
    record = get_object_or_404(BorrowRecord, pk=record_pk, user=request.user)
    
    if record.status == 'returned':
        messages.warning(request, 'This book has already been returned.')
        return redirect('borrow:my_books')
    
    if request.method == 'POST':
        # Process return
        record.return_book()
        
        # Update returned_to field using update to bypass validation
        BorrowRecord.objects.filter(pk=record.pk).update(returned_to=request.user)
        
        # Refresh from database
        record.refresh_from_db()
        
        messages.success(request, f'You have successfully returned "{record.book.title}".')
        
        # Show fine information if applicable
        if record.fine_amount > 0:
            if record.fine_paid:
                messages.info(request, f'Fine of ETB {record.fine_amount} has been paid.')
            else:
                messages.warning(request, f'You have an outstanding fine of ETB {record.fine_amount}. Please pay online.')
        
        return redirect('borrow:my_books')
    
    # Show confirmation page
    return render(request, 'borrow/student_return_confirm.html', {'record': record})


@login_required
def my_borrowed_books(request):
    """Show user's borrowed books"""
    user = request.user
    
    # Get current borrowed books
    borrowed_books = BorrowRecord.objects.filter(
        user=user, 
        status__in=['borrowed', 'overdue']
    ).select_related('book').order_by('-borrow_date')
    
    # Get pending requests
    pending_requests = BookRequest.objects.filter(
        user=user, 
        status='pending'
    ).select_related('book').order_by('-request_date')
    
    # Get ready requests waiting for pickup
    ready_requests = BookRequest.objects.filter(
        user=user, 
        status='ready'
    ).select_related('book').order_by('-approved_date')

    # Get recently rejected requests (last 14 days)
    from django.utils import timezone
    from datetime import timedelta
    two_weeks_ago = timezone.now() - timedelta(days=14)
    rejected_requests = BookRequest.objects.filter(
        user=user,
        status='rejected',
        updated_at__gte=two_weeks_ago
    ).select_related('book', 'approved_by').order_by('-updated_at')
    
    # Get cancelled requests (last 14 days)
    cancelled_requests = BookRequest.objects.filter(
        user=user,
        status='cancelled',
        updated_at__gte=two_weeks_ago
    ).select_related('book').order_by('-updated_at')
    
    # Get returned books (last 90 days) for rating and re-requesting
    ninety_days_ago = timezone.now() - timedelta(days=90)
    returned_books = BorrowRecord.objects.filter(
        user=user,
        status='returned',
        return_date__gte=ninety_days_ago
    ).select_related('book').order_by('-return_date')
    
    context = {
        'borrowed_books': borrowed_books,
        'pending_requests': pending_requests,
        'ready_requests': ready_requests,
        'rejected_requests': rejected_requests,
        'cancelled_requests': cancelled_requests,
        'returned_books': returned_books,
        'borrowed_count': borrowed_books.count(),
        'max_limit': user.profile.max_books_allowed
    }
    return render(request, 'borrow/my_borrowed_books.html', context)


@login_required
@student_required
def delete_request(request, request_pk):
    """Delete a single rejected or cancelled request"""
    book_request = get_object_or_404(BookRequest, pk=request_pk, user=request.user)
    
    # Only allow deletion of rejected or cancelled requests
    if book_request.status not in ['rejected', 'cancelled']:
        messages.error(request, 'Only rejected or cancelled requests can be deleted.')
        return redirect('borrow:my_books')
    
    if request.method == 'POST':
        book_title = book_request.book.title
        book_request.delete()
        messages.success(request, f'Request for "{book_title}" has been deleted.')
        return redirect('borrow:my_books')
    
    return redirect('borrow:my_books')


@login_required
@student_required
def clear_rejected_requests(request):
    """Clear all rejected requests for the current user"""
    if request.method == 'POST':
        deleted_count = BookRequest.objects.filter(
            user=request.user,
            status='rejected'
        ).delete()[0]
        
        if deleted_count > 0:
            messages.success(request, f'Successfully deleted {deleted_count} rejected request(s).')
        else:
            messages.info(request, 'No rejected requests to delete.')
        
        return redirect('borrow:my_books')
    
    return redirect('borrow:my_books')


@login_required
@student_required
def clear_cancelled_requests(request):
    """Clear all cancelled requests for the current user"""
    if request.method == 'POST':
        deleted_count = BookRequest.objects.filter(
            user=request.user,
            status='cancelled'
        ).delete()[0]
        
        if deleted_count > 0:
            messages.success(request, f'Successfully deleted {deleted_count} cancelled request(s).')
        else:
            messages.info(request, 'No cancelled requests to delete.')
        
        return redirect('borrow:my_books')
    
    return redirect('borrow:my_books')


@login_required
def borrow_history(request):
    """Student borrow history with search and filter"""
    user = request.user
    records = BorrowRecord.objects.filter(user=user).select_related('book').order_by('-borrow_date')

    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter in ['borrowed', 'returned', 'overdue']:
        records = records.filter(status=status_filter)

    # Search by book title
    search = request.GET.get('search', '').strip()
    if search:
        records = records.filter(book__title__icontains=search)

    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(records, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'records': page_obj.object_list,
        'status_filter': status_filter,
        'search': search,
        'total_count': records.count(),
    }
    return render(request, 'borrow/borrow_history.html', context)


@login_required
@user_passes_test(lambda user: user.is_librarian or user.is_admin)
def all_borrow_records(request):
    """All borrow records with search, filter and pagination"""
    from django.core.paginator import Paginator

    records = BorrowRecord.objects.select_related('user', 'book').order_by('-borrow_date')

    # Search
    search = request.GET.get('search', '').strip()
    if search:
        records = records.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(book__title__icontains=search)
        )

    # Status filter
    status_filter = request.GET.get('status', '')
    if status_filter in ['borrowed', 'returned', 'overdue']:
        records = records.filter(status=status_filter)

    total_count = records.count()
    paginator = Paginator(records, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'records': page_obj.object_list,
        'search': search,
        'status_filter': status_filter,
        'total_count': total_count,
    }
    return render(request, 'borrow/all_borrow_records.html', context)


@login_required
@librarian_or_admin_required
def overdue_management(request):
    """Overdue books with correct fine totals - shows both paid and unpaid"""
    from django.db.models import Sum
    from apps.dashboard.models import SystemSettings
    
    # Show ALL overdue books (both paid and unpaid fines)
    # Also include returned books that had overdue fines
    records = BorrowRecord.objects.filter(
        Q(status='overdue') | Q(status='returned', fine_amount__gt=0)
    ).select_related('user', 'book').order_by('-fine_paid', 'due_date')

    # Get system settings for currency conversion
    settings = SystemSettings.get_settings()
    
    # Calculate totals for currently overdue books only
    currently_overdue = BorrowRecord.objects.filter(status='overdue')
    total_fines = currently_overdue.aggregate(t=Sum('fine_amount'))['t'] or 0
    unpaid_fines = currently_overdue.filter(fine_paid=False).aggregate(t=Sum('fine_amount'))['t'] or 0
    unpaid_count = currently_overdue.filter(fine_paid=False).count()
    
    # For "collected" - ALL records that were overdue and are now paid (including returned ones)
    all_paid_fines = BorrowRecord.objects.filter(
        fine_amount__gt=0, 
        fine_paid=True
    ).aggregate(t=Sum('fine_amount'))['t'] or 0
    
    # Convert to USD
    total_fines_usd = float(total_fines) * float(settings.etb_to_usd_rate)
    unpaid_fines_usd = float(unpaid_fines) * float(settings.etb_to_usd_rate)
    paid_fines_usd = float(all_paid_fines) * float(settings.etb_to_usd_rate)

    context = {
        'records': records,
        'total_fines': total_fines,
        'unpaid_fines': unpaid_fines,
        'paid_fines': all_paid_fines,
        'unpaid_count': unpaid_count,
        'can_manage': request.user.is_admin,
        # USD amounts
        'total_fines_usd': total_fines_usd,
        'unpaid_fines_usd': unpaid_fines_usd,
        'paid_fines_usd': paid_fines_usd,
    }
    return render(request, 'borrow/overdue_management.html', context)


@login_required
@librarian_or_admin_required
def export_borrow_records_csv(request):
    """Export borrow records to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="borrow_records_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'User', 'Book Title', 'Author', 'ISBN', 'Borrow Date', 
        'Due Date', 'Return Date', 'Status', 'Fine Amount', 
        'Fine Paid', 'Issued By', 'Returned To'
    ])
    
    records = BorrowRecord.objects.select_related(
        'user', 'book', 'issued_by', 'returned_to'
    ).all()
    
    for record in records:
        writer.writerow([
            record.user.get_full_name() or record.user.username,
            record.book.title,
            record.book.author,
            record.book.isbn,
            record.borrow_date.strftime('%Y-%m-%d'),
            record.due_date.strftime('%Y-%m-%d'),
            record.return_date.strftime('%Y-%m-%d') if record.return_date else '',
            record.get_status_display(),
            record.fine_amount,
            'Yes' if record.fine_paid else 'No',
            record.issued_by.get_full_name() if record.issued_by else '',
            record.returned_to.get_full_name() if record.returned_to else ''
        ])
    
    return response


@login_required
@student_required
def request_list(request):
    """Show available books for students to request"""
    # Get ALL books (available and unavailable) so students see full catalog
    all_books = Book.objects.all().select_related('category')
    
    # Get user's current requests and borrows
    user_requests = BookRequest.objects.filter(
        user=request.user, 
        status__in=['pending', 'ready']
    ).values_list('book_id', flat=True)
    
    user_borrows = BorrowRecord.objects.filter(
        user=request.user, 
        status__in=['borrowed', 'overdue']
    ).values_list('book_id', flat=True)
    
    already_requested = set(user_requests)
    already_borrowed = set(user_borrows)
    
    # Get user's current borrow count
    current_borrowed = BorrowRecord.objects.filter(
        user=request.user, 
        status__in=['borrowed', 'overdue']
    ).count()
    
    # Check if user can borrow more books
    user_max_limit = request.user.profile.max_books_allowed
    can_request_more = current_borrowed < user_max_limit
    
    context = {
        'books': all_books,
        'already_requested': already_requested,
        'already_borrowed': already_borrowed,
        'current_borrowed': current_borrowed,
        'max_limit': user_max_limit,
        'can_request_more': can_request_more,
        'remaining_slots': user_max_limit - current_borrowed
    }
    return render(request, 'borrow/request_list.html', context)


@login_required
@student_required
def cancel_request(request, request_pk):
    """Student cancels their book request"""
    book_request = get_object_or_404(BookRequest, pk=request_pk, user=request.user)
    
    if not book_request.can_cancel():
        messages.error(request, f'Cannot cancel request. Current status: {book_request.get_status_display()}')
        return redirect('borrow:my_books')
    
    if request.method == 'POST':
        reason = request.POST.get('reason', 'Cancelled by user')
        if book_request.cancel(reason):
            messages.success(request, f'Your request for "{book_request.book.title}" has been cancelled.')
        else:
            messages.error(request, 'Unable to cancel request.')
        return redirect('borrow:my_books')
    
    return render(request, 'borrow/cancel_request.html', {'book_request': book_request})


@login_required
@student_required
def request_book(request, book_pk):
    """Student requests a book"""
    book = get_object_or_404(Book, pk=book_pk)
    user = request.user
    
    # Check for unpaid fines - BLOCK if any exist
    unpaid_fines = BorrowRecord.objects.filter(
        user=user,
        fine_amount__gt=0,
        fine_paid=False
    )
    
    if unpaid_fines.exists():
        total_unpaid = unpaid_fines.aggregate(total=Sum('fine_amount'))['total'] or 0
        messages.error(
            request, 
            f'You have unpaid fines totaling ETB {total_unpaid}. '
            f'Please pay your fines before requesting new books. '
            f'<a href="/borrow/my-books/" style="color: white; text-decoration: underline;">View and pay fines</a>',
            extra_tags='safe'
        )
        return redirect('books:detail', pk=book_pk)
    
    # Check if book is available
    if not book.is_available():
        messages.warning(request, f'All copies of "{book.title}" are currently borrowed. Your request has been added to the waitlist — you\'ll be notified when a copy becomes available.')
        # Still allow the request to be created as a waitlist entry
    
    # Check total limit: borrowed + pending/ready requests
    current_borrowed = BorrowRecord.objects.filter(
        user=user, 
        status__in=['borrowed', 'overdue']
    ).count()
    
    pending_requests = BookRequest.objects.filter(
        user=user,
        status__in=['pending', 'ready']
    ).count()
    
    total_books = current_borrowed + pending_requests
    user_max_limit = user.profile.max_books_allowed
    
    if total_books >= user_max_limit:
        messages.error(request, f'You have reached the maximum limit of {user_max_limit} books (currently borrowed: {current_borrowed}, pending requests: {pending_requests}). Please return a book or cancel a request before making a new request.')
        return redirect('borrow:request_list')
    
    # Check if already requested or borrowed
    existing_request = BookRequest.objects.filter(
        user=user, 
        book=book, 
        status__in=['pending', 'ready']
    ).exists()
    
    existing_borrow = BorrowRecord.objects.filter(
        user=user, 
        book=book, 
        status__in=['borrowed', 'overdue']
    ).exists()
    
    if existing_request:
        messages.warning(request, f'You have already requested "{book.title}".')
        return redirect('borrow:request_list')
    
    if existing_borrow:
        messages.warning(request, f'You have already borrowed "{book.title}".')
        return redirect('borrow:request_list')
    
    # Create book request with validation
    try:
        book_request = BookRequest(
            user=user,
            book=book,
            status='pending'
        )
        book_request.full_clean()  # Run model validation
        book_request.save()
    except ValidationError as e:
        # Extract error messages
        error_messages = []
        if hasattr(e, 'message_dict'):
            for field, errors in e.message_dict.items():
                error_messages.extend(errors)
        else:
            error_messages.append(str(e))
        
        for error_msg in error_messages:
            messages.error(request, error_msg)
        return redirect('borrow:request_list')
    
    messages.success(request, f'Your request for "{book.title}" has been submitted successfully.')
    return redirect('borrow:request_list')


@login_required
@librarian_or_admin_required
def pending_requests(request):
    """Show pending book requests with book availability info"""
    requests = BookRequest.objects.filter(status='pending').select_related('user', 'book', 'book__category').order_by('-request_date')

    context = {
        'requests': requests,
        'pending_count': requests.count()
    }
    return render(request, 'borrow/pending_requests.html', context)


@login_required
@librarian_or_admin_required
def approve_request(request, request_pk):
    """Librarian/Admin approves a book request and issues the book"""
    book_request = get_object_or_404(BookRequest, pk=request_pk)
    
    if book_request.status != 'pending':
        messages.error(request, 'This request has already been processed.')
        return redirect('borrow:pending_requests')
    
    # Check for unpaid fines - BLOCK approval if any exist
    unpaid_fines = BorrowRecord.objects.filter(
        user=book_request.user,
        fine_amount__gt=0,
        fine_paid=False
    )
    
    if unpaid_fines.exists():
        total_unpaid = unpaid_fines.aggregate(total=Sum('fine_amount'))['total'] or 0
        messages.error(
            request,
            f'{book_request.user.get_full_name() or book_request.user.username} has unpaid fines totaling ETB {total_unpaid}. '
            f'Cannot approve book request until fines are paid.'
        )
        return redirect('borrow:pending_requests')
    
    # Check if book is still available
    if not book_request.book.is_available():
        messages.error(request, f'"{book_request.book.title}" is no longer available.')
        return redirect('borrow:pending_requests')
    
    # Check user's total limit: borrowed + other pending/ready requests (excluding this one)
    current_borrowed = BorrowRecord.objects.filter(
        user=book_request.user, 
        status__in=['borrowed', 'overdue']
    ).count()
    
    # Count other pending/ready requests (excluding the current one being approved)
    other_requests = BookRequest.objects.filter(
        user=book_request.user,
        status__in=['pending', 'ready']
    ).exclude(pk=book_request.pk).count()
    
    total_books = current_borrowed + other_requests
    user_max_limit = book_request.user.profile.max_books_allowed
    
    # Check if approving this request would exceed the limit
    if total_books >= user_max_limit:
        messages.error(request, f'{book_request.user.username} has reached the maximum limit of {user_max_limit} books (currently borrowed: {current_borrowed}, other pending/ready requests: {other_requests}). Cannot approve this request.')
        return redirect('borrow:pending_requests')
    
    # Check for duplicate active borrow
    existing_borrow = BorrowRecord.objects.filter(
        user=book_request.user,
        book=book_request.book,
        status__in=['borrowed', 'overdue']
    ).exists()
    
    if existing_borrow:
        messages.error(request, f'{book_request.user.username} has already borrowed "{book_request.book.title}".')
        return redirect('borrow:pending_requests')
    
    # Approve request and create borrow record
    book_request.approve(request.user)
    
    # Create borrow record with validation
    try:
        due_date = timezone.now().date() + timezone.timedelta(days=settings.BORROW_DURATION_DAYS)
        borrow_record = BorrowRecord(
            user=book_request.user,
            book=book_request.book,
            book_request=book_request,
            due_date=due_date,
            issued_by=request.user
        )
        # Run validation
        borrow_record.full_clean()
        borrow_record.save()
        
        # Update book availability
        book_request.book.borrow_book()
        
        # Update user profile
        profile = book_request.user.profile
        profile.currently_borrowed += 1
        profile.save()
        
        # Mark request as fulfilled
        book_request.fulfill()
        
        # Send notification to user
        from apps.users.notifications import notify_request_approved
        notify_request_approved(request, book_request)
        
        # Clean up notification for all librarians/admins
        from apps.users.models import NotificationRead, User
        notification_key = f'pending_request_{book_request.pk}'
        librarians_and_admins = User.objects.filter(role__in=['librarian', 'admin'])
        for lib_user in librarians_and_admins:
            NotificationRead.objects.get_or_create(
                user=lib_user,
                notification_key=notification_key,
                notification_type='pending_request'
            )
        
        messages.success(request, f'Book "{book_request.book.title}" has been issued to {book_request.user.username}.')
        messages.info(request, f'Approval email sent to {book_request.user.email}.')
    except ValidationError as e:
        messages.error(request, f'Cannot issue book: {", ".join(e.messages)}')
    
    return redirect('borrow:pending_requests')


@login_required
@librarian_or_admin_required
def reject_request(request, request_pk):
    """Librarian/Admin rejects a book request"""
    book_request = get_object_or_404(BookRequest, pk=request_pk)
    
    if book_request.status != 'pending':
        messages.error(request, 'This request has already been processed.')
        return redirect('borrow:pending_requests')
    
    if request.method == 'POST':
        reason = request.POST.get('reason', 'No reason provided')
        book_request.reject(request.user, reason)
        
        # Send notification to user with the rejection reason
        from apps.users.notifications import notify_request_rejected
        notify_request_rejected(request, book_request, reason=book_request.rejection_reason)
        
        # Clean up notification for all librarians/admins
        from apps.users.models import NotificationRead, User
        notification_key = f'pending_request_{book_request.pk}'
        librarians_and_admins = User.objects.filter(role__in=['librarian', 'admin'])
        for lib_user in librarians_and_admins:
            NotificationRead.objects.get_or_create(
                user=lib_user,
                notification_key=notification_key,
                notification_type='pending_request'
            )
        
        messages.success(request, f'Request for "{book_request.book.title}" has been rejected.')
        messages.info(request, f'Rejection email sent to {book_request.user.email}.')
        return redirect('borrow:pending_requests')
    
    return render(request, 'borrow/reject_request.html', {'book_request': book_request})


@login_required
@user_passes_test(lambda user: user.is_librarian or user.is_admin)
def issue_return(request):
    """Issue & Return with search and filter"""
    from django.core.paginator import Paginator

    records = BorrowRecord.objects.filter(
        status__in=['borrowed', 'overdue']
    ).select_related('user', 'book').order_by('due_date')

    # Search
    search = request.GET.get('search', '').strip()
    if search:
        records = records.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(book__title__icontains=search)
        )

    # Status filter
    status_filter = request.GET.get('status', '')
    if status_filter in ['borrowed', 'overdue']:
        records = records.filter(status=status_filter)

    total_count = records.count()
    paginator = Paginator(records, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'records': page_obj.object_list,
        'search': search,
        'status_filter': status_filter,
        'total_count': total_count,
    }
    return render(request, 'borrow/issue_return.html', context)


@login_required
@user_passes_test(lambda user: user.is_librarian or user.is_admin)
def issue_book(request, book_pk, user_pk):
    return redirect('borrow:issue_return')



@login_required
@librarian_or_admin_required
def export_borrow_records_excel(request):
    """Export borrow records to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Borrow Records"
    
    # Define headers
    headers = [
        'User', 'Book Title', 'Author', 'ISBN', 'Borrow Date', 
        'Due Date', 'Return Date', 'Status', 'Fine Amount', 
        'Fine Paid', 'Issued By', 'Returned To'
    ]
    
    # Add headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add data
    records = BorrowRecord.objects.select_related(
        'user', 'book', 'issued_by', 'returned_to'
    ).all()
    
    for row, record in enumerate(records, 2):
        ws.cell(row=row, column=1, value=record.user.get_full_name() or record.user.username)
        ws.cell(row=row, column=2, value=record.book.title)
        ws.cell(row=row, column=3, value=record.book.author)
        ws.cell(row=row, column=4, value=record.book.isbn)
        ws.cell(row=row, column=5, value=record.borrow_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=6, value=record.due_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=7, value=record.return_date.strftime('%Y-%m-%d') if record.return_date else '')
        ws.cell(row=row, column=8, value=record.get_status_display())
        ws.cell(row=row, column=9, value=float(record.fine_amount))
        ws.cell(row=row, column=10, value='Yes' if record.fine_paid else 'No')
        ws.cell(row=row, column=11, value=record.issued_by.get_full_name() if record.issued_by else '')
        ws.cell(row=row, column=12, value=record.returned_to.get_full_name() if record.returned_to else '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="borrow_records_export.xlsx"'
    
    wb.save(response)
    return response