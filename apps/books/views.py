from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.utils import timezone
from .models import Book, Category, BookReview
from .forms import BookForm, CategoryForm, BookSearchForm, BookReviewForm
from django.core.paginator import Paginator
from ..users.permissions import librarian_or_admin_required, admin_required
from ..borrow.models import BorrowRecord
import csv
import os
import logging

logger = logging.getLogger(__name__)


def check_librarian_or_admin(user):
    """Check if user is librarian or admin"""
    return user.is_librarian or user.is_admin


def books_list(request):
    """List all books with search and filtering"""
    form = BookSearchForm(request.GET or None)
    books = Book.objects.select_related('category').all()
    
    if form.is_valid():
        search = form.cleaned_data.get('search')
        category = form.cleaned_data.get('category')
        language = form.cleaned_data.get('language')
        availability = form.cleaned_data.get('availability')
        
        if search:
            books = books.filter(
                Q(title__icontains=search) |
                Q(author__icontains=search) |
                Q(isbn__icontains=search)
            )
        
        if category:
            books = books.filter(category=category)
        
        if language:
            books = books.filter(language=language)
        
        if availability == 'available':
            books = books.filter(available_copies__gt=0)
        elif availability == 'unavailable':
            books = books.filter(available_copies=0)
    
    # Pagination
    paginator = Paginator(books, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'books': page_obj.object_list,
        'form': form,
    }
    
    return render(request, 'books/books_list.html', context)


def book_detail(request, pk):
    """Book detail view with reviews, QR code, and PDF reading"""
    book = get_object_or_404(Book, pk=pk)
    reviews = book.reviews.all().select_related('user')
    
    # Calculate rating distribution
    rating_distribution = {5: 0, 4: 0, 3: 0, 2: 0, 1: 0}
    for review in reviews:
        rating_distribution[review.rating] += 1
    
    # Calculate average rating - always sync with actual reviews
    from django.db.models import Avg
    if reviews.exists():
        avg_rating = reviews.aggregate(Avg('rating'))['rating__avg']
        correct_rating = round(avg_rating, 2) if avg_rating else 0.00
    else:
        correct_rating = 0.00

    # Update DB if rating is out of sync
    if book.rating != correct_rating:
        book.rating = correct_rating
        book.save(update_fields=['rating'])
        book.refresh_from_db()

    # Check if user can review (has borrowed and returned)
    can_review = False
    has_borrowed_and_returned = False
    active_borrow_record = None
    
    if request.user.is_authenticated:
        has_borrowed_and_returned = BorrowRecord.objects.filter(
            user=request.user,
            book=book,
            status='returned'
        ).exists()
        
        has_reviewed = book.reviews.filter(user=request.user).exists()
        can_review = has_borrowed_and_returned and not has_reviewed
        
        # Check if user currently has this book borrowed
        active_borrow_record = BorrowRecord.objects.filter(
            user=request.user,
            book=book,
            status__in=['borrowed', 'overdue']
        ).first()
    
    # Generate QR code if missing (e.g. legacy books or failed prior save)
    if not book.qr_code:
        try:
            book.generate_qr_code()
            book.save(update_fields=['qr_code'])
        except Exception as e:
            logger.warning('Could not generate QR code for book %s: %s', book.pk, e)
    
    context = {
        'book': book,
        'reviews': reviews,
        'has_reviewed': request.user.is_authenticated and book.reviews.filter(user=request.user).exists(),
        'can_review': can_review,
        'has_borrowed_and_returned': has_borrowed_and_returned,
        'active_borrow_record': active_borrow_record,
        'rating_distribution': rating_distribution,
        'total_reviews': reviews.count(),
    }
    
    return render(request, 'books/book_detail.html', context)


@login_required
@librarian_or_admin_required
def book_create(request):
    """Create a new book (Librarian/Admin only)"""
    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES)
        if form.is_valid():
            book = form.save(commit=False)
            # Set available copies equal to total copies for new books
            book.available_copies = book.total_copies
            book.save()
            
            # Log activity
            from apps.dashboard.utils import log_activity
            log_activity(request.user, 'book_added', f'Added book: "{book.title}" (ISBN: {book.isbn})', request)
            
            messages.success(request, f'Book "{book.title}" has been added successfully.')
            return redirect('books:detail', pk=book.pk)
    else:
        form = BookForm()
    
    return render(request, 'books/book_form.html', {'form': form, 'action': 'Add'})


@login_required
@admin_required
def book_edit(request, pk):
    """Edit a book (Admin only)"""
    book = get_object_or_404(Book, pk=pk)
    
    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES, instance=book)
        if form.is_valid():
            form.save()
            
            # Log activity
            from apps.dashboard.utils import log_activity
            log_activity(request.user, 'book_updated', f'Updated book: "{book.title}" (ISBN: {book.isbn})', request)
            
            messages.success(request, f'Book "{book.title}" has been updated.')
            return redirect('books:detail', pk=book.pk)
    else:
        form = BookForm(instance=book)
    
    return render(request, 'books/book_form.html', {'form': form, 'book': book, 'action': 'Edit'})


@login_required
@admin_required
def book_delete(request, pk):
    """Delete a book (Admin only)"""
    book = get_object_or_404(Book, pk=pk)
    
    if request.method == 'POST':
        title = book.title
        isbn = book.isbn
        
        # Check for active borrows
        from apps.borrow.models import BorrowRecord
        active_borrows = BorrowRecord.objects.filter(
            book=book, 
            status__in=['borrowed', 'overdue']
        ).count()
        
        if active_borrows > 0:
            messages.warning(
                request, 
                f'Warning: "{title}" has {active_borrows} active borrow(s). '
                f'Deleting this book will affect those records.'
            )
        
        # Log activity before deletion
        from apps.dashboard.utils import log_activity
        log_activity(request.user, 'book_deleted', f'Deleted book: "{title}" (ISBN: {isbn})', request)
        
        # Delete the book (this will cascade delete related records)
        book.delete()
        
        messages.success(request, f'Book "{title}" has been permanently deleted from the library.')
        return redirect('books:list')
    
    # If GET request, show confirmation page (though we're using modal now)
    return render(request, 'books/book_confirm_delete.html', {'book': book})


@login_required
def add_review(request, book_pk):
    """Add a review to a book - Students only, must have borrowed and returned"""
    book = get_object_or_404(Book, pk=book_pk)
    
    # Only students can review
    if not request.user.is_student:
        messages.error(request, 'Only students can write reviews.')
        return redirect('books:detail', pk=book.pk)
    
    # Check if user already reviewed this book
    existing_review = BookReview.objects.filter(book=book, user=request.user).first()
    
    # Check if user has borrowed and returned this book
    has_borrowed = BorrowRecord.objects.filter(
        user=request.user,
        book=book,
        status='returned'
    ).exists()
    
    # Enforce borrow requirement for students
    if not has_borrowed:
        messages.error(request, 'You can only review books you have borrowed and returned.')
        return redirect('books:detail', pk=book.pk)
    
    if request.method == 'POST':
        form = BookReviewForm(request.POST, instance=existing_review)
        if form.is_valid():
            review = form.save(commit=False)
            review.book = book
            review.user = request.user
            
            try:
                # Students always go through validation
                review.save()
                
                # Update book's average rating
                from django.db.models import Avg
                avg_rating = book.reviews.aggregate(Avg('rating'))['rating__avg']
                if avg_rating:
                    book.rating = round(avg_rating, 2)
                    book.save(update_fields=['rating'])
                
                # Log activity
                from apps.dashboard.utils import log_activity
                action = 'review_updated' if existing_review else 'review_added'
                log_activity(request.user, action, f'{"Updated" if existing_review else "Added"} review for "{book.title}"', request)
                
                if existing_review:
                    messages.success(request, 'Your review has been updated successfully.')
                else:
                    messages.success(request, 'Your review has been added successfully.')
                return redirect('books:detail', pk=book.pk)
            except Exception as e:
                messages.error(request, f'Error saving review: {str(e)}')
    else:
        form = BookReviewForm(instance=existing_review)
    
    context = {
        'form': form,
        'book': book,
        'existing_review': existing_review,
        'has_borrowed': has_borrowed
    }
    
    return render(request, 'books/add_review.html', context)


@login_required
@admin_required
def delete_review(request, review_pk):
    """Delete a review (Admin only) - for removing inappropriate content"""
    review = get_object_or_404(BookReview, pk=review_pk)
    book = review.book
    
    if request.method == 'POST':
        reviewer_name = review.user.get_full_name() or review.user.username
        
        # Log activity before deletion
        from apps.dashboard.utils import log_activity
        log_activity(
            request.user, 
            'review_deleted', 
            f'Deleted review by {reviewer_name} for "{book.title}"', 
            request
        )
        
        # Delete the review
        review.delete()
        
        # Update book's average rating
        from django.db.models import Avg
        remaining_reviews = book.reviews.all()
        if remaining_reviews.exists():
            avg_rating = remaining_reviews.aggregate(Avg('rating'))['rating__avg']
            book.rating = round(avg_rating, 2)
        else:
            book.rating = 0.00
        book.save(update_fields=['rating'])
        
        messages.success(request, f'Review by {reviewer_name} has been deleted.')
        return redirect('books:detail', pk=book.pk)
    
    # If GET request, show confirmation (though we'll use modal)
    context = {
        'review': review,
        'book': book
    }
    return render(request, 'books/review_confirm_delete.html', context)


def categories_list(request):
    """List all categories"""
    categories = Category.objects.all()
    
    context = {
        'categories': categories,
    }
    
    return render(request, 'books/categories_list.html', context)


def category_detail(request, pk):
    """Category detail with books"""
    category = get_object_or_404(Category, pk=pk)
    books = category.books.all()
    
    context = {
        'category': category,
        'books': books,
    }
    
    return render(request, 'books/category_detail.html', context)


@login_required
@user_passes_test(check_librarian_or_admin)
def category_create(request):
    """Create a new category (Librarian/Admin only)"""
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category "{category.name}" has been created.')
            return redirect('books:categories')
    else:
        form = CategoryForm()
    
    return render(request, 'books/category_form.html', {'form': form, 'action': 'Add'})


@login_required
@user_passes_test(check_librarian_or_admin)
def category_edit(request, pk):
    """Edit a category (Librarian/Admin only)"""
    category = get_object_or_404(Category, pk=pk)
    
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, f'Category "{category.name}" has been updated.')
            return redirect('books:category_detail', pk=category.pk)
    else:
        form = CategoryForm(instance=category)
    
    return render(request, 'books/category_form.html', {'form': form, 'category': category, 'action': 'Edit'})


@login_required
@admin_required
def export_books_csv(request):
    """Export books to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="books_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'ISBN', 'Title', 'Author', 'Category', 'Total Copies', 
        'Available Copies', 'Publisher', 'Language', 'Times Borrowed',
        'Publication Date', 'Pages', 'Created At'
    ])
    
    for book in Book.objects.select_related('category').all():
        writer.writerow([
            book.isbn,
            book.title,
            book.author,
            book.category.name if book.category else 'Uncategorized',
            book.total_copies,
            book.available_copies,
            book.publisher,
            book.get_language_display(),
            book.times_borrowed,
            book.publication_date,
            book.pages,
            book.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    return response


@login_required
def read_pdf(request, pk):
    """PDF reader view for books"""
    book = get_object_or_404(Book, pk=pk)
    
    # Check if user has access to read this book
    user = request.user
    can_read = False
    
    # Admin and librarians can read any book
    if user.is_admin or user.is_librarian:
        can_read = True
    # Students can read if they have borrowed the book or if it's freely available
    elif user.is_student:
        # Check if user has borrowed this book
        has_borrowed = user.borrow_records.filter(
            book=book, 
            status__in=['borrowed', 'overdue']
        ).exists()
        can_read = has_borrowed
    
    if not can_read:
        messages.error(request, "You don't have permission to read this book. Please borrow it first.")
        return redirect('books:detail', pk=pk)
    
    # Check if PDF exists
    if not book.pdf_file:
        messages.warning(request, f'PDF file is not available for "{book.title}". The digital version has not been uploaded yet.')
        return redirect('books:detail', pk=pk)

    # Check if the file actually exists in storage (skip for Cloudinary — .exists() not reliable)
    storage = book.pdf_file.storage
    is_cloudinary = 'cloudinary' in type(storage).__module__.lower()

    if not is_cloudinary:
        try:
            if not storage.exists(book.pdf_file.name):
                messages.error(request, f'PDF file for "{book.title}" is missing from the server. Please contact the librarian.')
                return redirect('books:detail', pk=pk)
        except Exception as e:
            messages.error(request, f'Error accessing PDF file: {str(e)}')
            return redirect('books:detail', pk=pk)
    
    context = {
        'book': book,
        'pdf_url': book.pdf_file.url
    }
    return render(request, 'books/pdf_reader.html', context)

@login_required
def serve_pdf(request, pk):
    """Serve PDF file with access control.
    
    On Cloudinary (production): redirects to the Cloudinary URL.
    On local storage (development): reads and serves the file from disk.
    """
    book = get_object_or_404(Book, pk=pk)

    # Access control
    user = request.user
    can_read = False

    if user.is_admin or user.is_librarian:
        can_read = True
    elif user.is_student:
        has_borrowed = user.borrow_records.filter(
            book=book,
            status__in=['borrowed', 'overdue']
        ).exists()
        can_read = has_borrowed

    if not can_read:
        logger.warning(
            f"PDF access denied for user {user.username} (ID: {user.id}) "
            f"attempting to access book '{book.title}' (ID: {book.id})"
        )
        return HttpResponse("Access denied. You must borrow this book first.", status=403)

    if not book.pdf_file:
        logger.error(
            f"PDF file not configured for book '{book.title}' (ID: {book.id}). "
            f"Requested by user {user.username} (ID: {user.id})"
        )
        return HttpResponse("PDF file not available for this book.", status=404)

    try:
        # Detect if we are using Cloudinary storage (no local .path support)
        storage = book.pdf_file.storage
        is_cloudinary = 'cloudinary' in type(storage).__module__.lower()

        if is_cloudinary:
            # On Cloudinary: redirect directly to the remote URL
            pdf_url = book.pdf_file.url
            logger.info(
                f"PDF redirected to Cloudinary URL for book '{book.title}' (ID: {book.id}) "
                f"to user {user.username} (ID: {user.id})"
            )
            from django.http import HttpResponseRedirect
            return HttpResponseRedirect(pdf_url)

        # Local storage: check file exists then serve from disk
        if not storage.exists(book.pdf_file.name):
            logger.error(
                f"PDF file missing from local storage for book '{book.title}' (ID: {book.id}). "
                f"Expected path: {book.pdf_file.name}"
            )
            return HttpResponse("PDF file is missing from the server.", status=404)

        file_path = book.pdf_file.path
        file_size = os.path.getsize(file_path)

        with open(file_path, 'rb') as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type='application/pdf')

        response['Content-Disposition'] = f'inline; filename="{book.title}.pdf"'
        response['Content-Length'] = file_size
        response['X-Frame-Options'] = 'SAMEORIGIN'

        logger.info(
            f"PDF served from disk for book '{book.title}' (ID: {book.id}) "
            f"to user {user.username} (ID: {user.id}). Size: {file_size} bytes"
        )
        return response

    except Exception as e:
        logger.exception(
            f"Unexpected error serving PDF for book '{book.title}' (ID: {book.id}): {str(e)}. "
            f"User: {user.username} (ID: {user.id})"
        )
        return HttpResponse("An unexpected error occurred while accessing the PDF.", status=500)


def get_book_recommendations(request, pk):
    """Get book recommendations based on category and popularity"""
    book = get_object_or_404(Book, pk=pk)
    
    # Get books from same category
    category_books = Book.objects.filter(
        category=book.category,
        available_copies__gt=0
    ).exclude(pk=book.pk).order_by('-times_borrowed')[:5]
    
    # Get most popular books if category doesn't have enough
    if category_books.count() < 5:
        popular_books = Book.objects.filter(
            available_copies__gt=0
        ).exclude(pk=book.pk).order_by('-times_borrowed')[:5]
        
        # Combine and remove duplicates
        recommendations = list(category_books) + [
            b for b in popular_books if b not in category_books
        ]
        recommendations = recommendations[:5]
    else:
        recommendations = list(category_books)
    
    recommendations_data = [{
        'id': b.id,
        'title': b.title,
        'author': b.author,
        'cover_url': b.cover_image.url if b.cover_image else None,
        'times_borrowed': b.times_borrowed
    } for b in recommendations]
    
    return JsonResponse({'recommendations': recommendations_data})

@login_required
@librarian_or_admin_required
def manage_stock(request):
    """Manage book stock - adjust total and available copies"""
    from django.db.models import Sum, F, ExpressionWrapper, IntegerField

    books = Book.objects.select_related('category').annotate(
        borrowed_copies=ExpressionWrapper(
            F('total_copies') - F('available_copies'),
            output_field=IntegerField()
        )
    ).order_by('title')

    if request.method == 'POST':
        book_id = request.POST.get('book_id')
        action = request.POST.get('action')
        amount = int(request.POST.get('amount', 1))
        book = get_object_or_404(Book, pk=book_id)

        if action == 'add_copies':
            book.total_copies += amount
            book.available_copies += amount
            book.save()
            messages.success(request, f'Added {amount} cop{"y" if amount == 1 else "ies"} to "{book.title}". Total: {book.total_copies}, Available: {book.available_copies}.')

        elif action == 'remove_copies':
            if amount > book.available_copies:
                messages.error(request, f'Cannot remove {amount} copies — only {book.available_copies} available (not currently borrowed).')
            elif book.total_copies - amount < 1:
                messages.error(request, f'Cannot remove {amount} copies — must keep at least 1 total copy.')
            else:
                book.total_copies -= amount
                book.available_copies -= amount
                book.save()
                messages.success(request, f'Removed {amount} cop{"y" if amount == 1 else "ies"} from "{book.title}". Total: {book.total_copies}, Available: {book.available_copies}.')

        elif action == 'set_available':
            currently_borrowed = book.total_copies - book.available_copies
            new_available = int(request.POST.get('amount', book.available_copies))
            if new_available < 0:
                messages.error(request, 'Available copies cannot be negative.')
            elif new_available > book.total_copies - currently_borrowed:
                messages.error(request, f'Cannot set available to {new_available} — {currently_borrowed} copies are currently borrowed.')
            else:
                book.available_copies = new_available
                book.save()
                messages.success(request, f'Updated available copies for "{book.title}" to {new_available}.')

        return redirect('books:manage_stock')

    # Summary stats
    from django.db.models import Sum
    total_books = books.count()
    total_copies = books.aggregate(t=Sum('total_copies'))['t'] or 0
    total_available = books.aggregate(a=Sum('available_copies'))['a'] or 0
    total_borrowed = total_copies - total_available
    out_of_stock = books.filter(available_copies=0).count()
    low_stock = books.filter(available_copies__gt=0, available_copies__lte=2).count()

    context = {
        'books': books,
        'total_books': total_books,
        'total_copies': total_copies,
        'total_available': total_available,
        'total_borrowed': total_borrowed,
        'out_of_stock': out_of_stock,
        'low_stock': low_stock,
    }
    return render(request, 'books/manage_stock.html', context)


@login_required
def book_recommendations(request):
    """Get book recommendations for the current user"""
    user = request.user
    
    # Get user's borrow history to understand preferences
    user_borrows = BorrowRecord.objects.filter(user=user).select_related('book__category')
    
    # Get categories user has borrowed from
    user_categories = user_borrows.values_list('book__category', flat=True).distinct()
    
    # Get books from same categories (excluding already borrowed)
    borrowed_book_ids = user_borrows.values_list('book_id', flat=True)
    
    recommended_books = Book.objects.filter(
        category__in=user_categories,
        available_copies__gt=0
    ).exclude(
        id__in=borrowed_book_ids
    ).order_by('-times_borrowed', '-rating')[:10]
    
    # If no category-based recommendations, show most popular available books
    if not recommended_books.exists():
        recommended_books = Book.objects.filter(
            available_copies__gt=0
        ).exclude(
            id__in=borrowed_book_ids
        ).order_by('-times_borrowed', '-rating')[:10]
    
    context = {
        'recommended_books': recommended_books,
        'user_categories': Category.objects.filter(id__in=user_categories),
    }
    
    return render(request, 'books/recommendations.html', context)
    

# ============================================================================
# BULK OPERATIONS
# ============================================================================

@login_required
@admin_required
def bulk_import_books(request):
    """Bulk import books from CSV/Excel file with optional cover images and PDFs"""
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        zip_file = request.FILES.get('zip_file')
        
        if not uploaded_file:
            messages.error(request, 'Please select a CSV/Excel file to upload.')
            return redirect('books:bulk_import')
        
        # Check file extension
        file_extension = uploaded_file.name.split('.')[-1].lower()
        
        if file_extension not in ['csv', 'xlsx', 'xls']:
            messages.error(request, 'Invalid file format. Please upload CSV or Excel file.')
            return redirect('books:bulk_import')
        
        try:
            import zipfile
            import tempfile
            import shutil
            from pathlib import Path
            
            imported_count = 0
            skipped_count = 0
            errors = []
            files_attached = {'covers': 0, 'pdfs': 0}
            files_missing = {'covers': 0, 'pdfs': 0}
            
            # Extract ZIP file if provided
            temp_dir = None
            covers_dict = {}
            pdfs_dict = {}
            
            if zip_file:
                try:
                    temp_dir = tempfile.mkdtemp()
                    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                        zip_ref.extractall(temp_dir)
                    
                    # Build dictionaries of available files
                    for root, dirs, files in os.walk(temp_dir):
                        for filename in files:
                            file_path = os.path.join(root, filename)
                            file_lower = filename.lower()
                            
                            # Check if it's in covers folder or has image extension
                            if 'cover' in root.lower() or file_lower.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
                                covers_dict[filename.lower()] = file_path
                            
                            # Check if it's in pdfs folder or has pdf extension
                            if 'pdf' in root.lower() or file_lower.endswith('.pdf'):
                                pdfs_dict[filename.lower()] = file_path
                    
                except Exception as e:
                    messages.warning(request, f'Could not extract ZIP file: {str(e)}. Continuing without files.')
            
            imported_count = 0
            skipped_count = 0
            errors = []
            
            if file_extension == 'csv':
                # Process CSV file
                import csv
                import io
                
                decoded_file = uploaded_file.read().decode('utf-8')
                csv_reader = csv.DictReader(io.StringIO(decoded_file))
                
                for row_num, row in enumerate(csv_reader, start=2):
                    try:
                        # Get or create category
                        category_name = row.get('category', '').strip()
                        category = None
                        if category_name:
                            category, _ = Category.objects.get_or_create(
                                name=category_name,
                                defaults={'description': f'Category for {category_name}'}
                            )
                        
                        # Check if book already exists
                        isbn = row.get('isbn', '').strip()
                        if isbn and Book.objects.filter(isbn=isbn).exists():
                            skipped_count += 1
                            errors.append(f"Row {row_num}: Book with ISBN {isbn} already exists")
                            continue
                        
                        # Create book
                        book = Book.objects.create(
                            isbn=isbn or f'AUTO-{timezone.now().timestamp()}',
                            title=row.get('title', '').strip(),
                            author=row.get('author', '').strip(),
                            description=row.get('description', '').strip(),
                            category=category,
                            total_copies=int(row.get('total_copies', 1)),
                            available_copies=int(row.get('available_copies', row.get('total_copies', 1))),
                            publisher=row.get('publisher', '').strip(),
                            publication_date=row.get('publication_date', None) or None,
                            pages=int(row.get('pages', 0)) if row.get('pages') else None,
                            language=row.get('language', 'en').strip()[:2],
                        )
                        
                        # Try to attach cover image (optional)
                        cover_filename = row.get('cover_filename', '').strip()
                        if cover_filename and cover_filename.lower() in covers_dict:
                            try:
                                with open(covers_dict[cover_filename.lower()], 'rb') as f:
                                    from django.core.files import File
                                    book.cover_image.save(cover_filename, File(f), save=True)
                                files_attached['covers'] += 1
                            except Exception as e:
                                files_missing['covers'] += 1
                                errors.append(f"Row {row_num}: Could not attach cover '{cover_filename}': {str(e)}")
                        elif cover_filename:
                            files_missing['covers'] += 1
                        
                        # Try to attach PDF (optional)
                        pdf_filename = row.get('pdf_filename', '').strip()
                        if pdf_filename and pdf_filename.lower() in pdfs_dict:
                            try:
                                with open(pdfs_dict[pdf_filename.lower()], 'rb') as f:
                                    from django.core.files import File
                                    book.pdf_file.save(pdf_filename, File(f), save=True)
                                files_attached['pdfs'] += 1
                            except Exception as e:
                                files_missing['pdfs'] += 1
                                errors.append(f"Row {row_num}: Could not attach PDF '{pdf_filename}': {str(e)}")
                        elif pdf_filename:
                            files_missing['pdfs'] += 1
                        
                        imported_count += 1
                        
                    except Exception as e:
                        skipped_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")
            
            else:
                # Process Excel file
                import openpyxl
                
                workbook = openpyxl.load_workbook(uploaded_file)
                sheet = workbook.active
                
                # Get headers from first row
                headers = [cell.value for cell in sheet[1]]
                
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Create dictionary from headers and row values
                        row_dict = dict(zip(headers, row))
                        
                        # Get or create category
                        category_name = str(row_dict.get('category', '')).strip()
                        category = None
                        if category_name:
                            category, _ = Category.objects.get_or_create(
                                name=category_name,
                                defaults={'description': f'Category for {category_name}'}
                            )
                        
                        # Check if book already exists
                        isbn = str(row_dict.get('isbn', '')).strip()
                        if isbn and Book.objects.filter(isbn=isbn).exists():
                            skipped_count += 1
                            errors.append(f"Row {row_num}: Book with ISBN {isbn} already exists")
                            continue
                        
                        # Create book
                        book = Book.objects.create(
                            isbn=isbn or f'AUTO-{timezone.now().timestamp()}',
                            title=str(row_dict.get('title', '')).strip(),
                            author=str(row_dict.get('author', '')).strip(),
                            description=str(row_dict.get('description', '')).strip(),
                            category=category,
                            total_copies=int(row_dict.get('total_copies', 1)),
                            available_copies=int(row_dict.get('available_copies', row_dict.get('total_copies', 1))),
                            publisher=str(row_dict.get('publisher', '')).strip(),
                            publication_date=row_dict.get('publication_date', None) or None,
                            pages=int(row_dict.get('pages', 0)) if row_dict.get('pages') else None,
                            language=str(row_dict.get('language', 'en')).strip()[:2],
                        )
                        
                        # Try to attach cover image (optional)
                        cover_filename = str(row_dict.get('cover_filename', '')).strip()
                        if cover_filename and cover_filename.lower() in covers_dict:
                            try:
                                with open(covers_dict[cover_filename.lower()], 'rb') as f:
                                    from django.core.files import File
                                    book.cover_image.save(cover_filename, File(f), save=True)
                                files_attached['covers'] += 1
                            except Exception as e:
                                files_missing['covers'] += 1
                                errors.append(f"Row {row_num}: Could not attach cover '{cover_filename}': {str(e)}")
                        elif cover_filename:
                            files_missing['covers'] += 1
                        
                        # Try to attach PDF (optional)
                        pdf_filename = str(row_dict.get('pdf_filename', '')).strip()
                        if pdf_filename and pdf_filename.lower() in pdfs_dict:
                            try:
                                with open(pdfs_dict[pdf_filename.lower()], 'rb') as f:
                                    from django.core.files import File
                                    book.pdf_file.save(pdf_filename, File(f), save=True)
                                files_attached['pdfs'] += 1
                            except Exception as e:
                                files_missing['pdfs'] += 1
                                errors.append(f"Row {row_num}: Could not attach PDF '{pdf_filename}': {str(e)}")
                        elif pdf_filename:
                            files_missing['pdfs'] += 1
                        
                        imported_count += 1
                        
                    except Exception as e:
                        skipped_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")
            
            # Clean up temp directory
            if temp_dir and os.path.exists(temp_dir):
                try:
                    shutil.rmtree(temp_dir)
                except:
                    pass
            
            # Log activity
            from apps.dashboard.utils import log_activity
            log_activity(
                request.user, 
                'book_added', 
                f'Bulk imported {imported_count} books from {uploaded_file.name}', 
                request
            )
            
            # Show results
            if imported_count > 0:
                messages.success(request, f'Successfully imported {imported_count} book(s).')
            
            if files_attached['covers'] > 0 or files_attached['pdfs'] > 0:
                messages.success(request, f"Attached {files_attached['covers']} cover(s) and {files_attached['pdfs']} PDF(s).")
            
            if files_missing['covers'] > 0 or files_missing['pdfs'] > 0:
                messages.info(request, f"Could not find {files_missing['covers']} cover(s) and {files_missing['pdfs']} PDF(s) in ZIP file.")
            
            if skipped_count > 0:
                messages.warning(request, f'Skipped {skipped_count} book(s). Check errors below.')
            
            # Store errors in session for display
            if errors:
                request.session['bulk_import_errors'] = errors[:50]  # Limit to 50 errors
            
        except Exception as e:
            logger.exception(f"Error during bulk import: {str(e)}")
            messages.error(request, f'Error processing file: {str(e)}')
        
        return redirect('books:bulk_import')
    
    # GET request - show upload form
    errors = request.session.pop('bulk_import_errors', [])
    
    context = {
        'errors': errors,
    }
    return render(request, 'books/bulk_import.html', context)

@login_required
@admin_required
def manage_books(request):
    """Manage Books page — Admin only. Edit, delete, bulk delete, bulk stock increase."""
    from django.db.models import Sum, F, ExpressionWrapper, IntegerField

    # ── Bulk actions (POST) ──────────────────────────────────────────────────
    if request.method == 'POST':
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_books')

        if action == 'bulk_delete':
            if not selected_ids:
                messages.warning(request, 'No books selected for deletion.')
                return redirect('books:manage_books')
            books_qs = Book.objects.filter(pk__in=selected_ids)
            count = books_qs.count()
            # Check for active borrows
            from apps.borrow.models import BorrowRecord
            active = BorrowRecord.objects.filter(
                book__in=books_qs, status__in=['borrowed', 'overdue']
            ).count()
            if active > 0:
                messages.warning(
                    request,
                    f'{active} active borrow(s) exist for selected books. '
                    f'Deleting anyway — borrow records will be affected.'
                )
            from apps.dashboard.utils import log_activity
            log_activity(request.user, 'book_deleted',
                         f'Bulk deleted {count} book(s) from Manage Books', request)
            books_qs.delete()
            messages.success(request, f'Successfully deleted {count} book(s).')

        elif action == 'bulk_add_copies':
            if not selected_ids:
                messages.warning(request, 'No books selected.')
                return redirect('books:manage_books')
            try:
                amount = int(request.POST.get('copies_amount', 1))
                if amount < 1:
                    raise ValueError
            except (ValueError, TypeError):
                messages.error(request, 'Invalid copies amount.')
                return redirect('books:manage_books')
            books_qs = Book.objects.filter(pk__in=selected_ids)
            count = books_qs.count()
            # Use F() expressions with queryset update to bypass model save/full_clean
            books_qs.update(
                total_copies=F('total_copies') + amount,
                available_copies=F('available_copies') + amount
            )
            from apps.dashboard.utils import log_activity
            log_activity(request.user, 'book_updated',
                         f'Bulk added {amount} cop(ies) to {count} book(s)', request)
            messages.success(
                request,
                f'Added {amount} cop{"y" if amount == 1 else "ies"} to {count} book(s).'
            )

        return redirect('books:manage_books')

    # ── GET — build queryset with filters ───────────────────────────────────
    books = Book.objects.select_related('category').annotate(
        borrowed_copies=ExpressionWrapper(
            F('total_copies') - F('available_copies'),
            output_field=IntegerField()
        )
    )

    # Search
    search = request.GET.get('search', '').strip()
    if search:
        books = books.filter(
            Q(title__icontains=search) |
            Q(author__icontains=search) |
            Q(isbn__icontains=search)
        )

    # Category filter
    category_id = request.GET.get('category', '')
    if category_id:
        books = books.filter(category_id=category_id)

    # Availability filter
    availability = request.GET.get('availability', '')
    if availability == 'available':
        books = books.filter(available_copies__gt=0)
    elif availability == 'unavailable':
        books = books.filter(available_copies=0)
    elif availability == 'low_stock':
        books = books.filter(available_copies__gt=0, available_copies__lte=2)

    # Sort
    sort = request.GET.get('sort', '-created_at')
    allowed_sorts = ['title', '-title', 'author', '-author', '-created_at',
                     'created_at', '-times_borrowed', 'available_copies', '-available_copies']
    if sort not in allowed_sorts:
        sort = '-created_at'
    books = books.order_by(sort)

    # Pagination
    paginator = Paginator(books, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Summary stats
    all_books = Book.objects.aggregate(
        total=Count('id'),
        total_copies=Sum('total_copies'),
        available=Sum('available_copies'),
    )
    total_books   = all_books['total'] or 0
    total_copies  = all_books['total_copies'] or 0
    total_avail   = all_books['available'] or 0
    total_borrowed = total_copies - total_avail
    out_of_stock  = Book.objects.filter(available_copies=0).count()
    low_stock     = Book.objects.filter(available_copies__gt=0, available_copies__lte=2).count()

    categories = Category.objects.all()

    context = {
        'page_obj': page_obj,
        'books': page_obj.object_list,
        'categories': categories,
        'search': search,
        'selected_category': category_id,
        'selected_availability': availability,
        'selected_sort': sort,
        # Stats
        'total_books': total_books,
        'total_copies': total_copies,
        'total_available': total_avail,
        'total_borrowed': total_borrowed,
        'out_of_stock': out_of_stock,
        'low_stock': low_stock,
    }
    return render(request, 'books/manage_books.html', context)


@login_required
@admin_required
def download_import_template(request):
    """Download CSV or Excel template for bulk book import"""
    format_type = request.GET.get('format', 'csv').lower()
    
    # Template data
    headers = [
        'isbn', 'title', 'author', 'description', 'category', 
        'total_copies', 'available_copies', 'publisher', 
        'publication_date', 'pages', 'language', 'cover_filename', 'pdf_filename'
    ]
    
    example_data = [
        [
            '978-0-123456-78-9',
            'Example Book Title',
            'John Doe',
            'This is a sample book description',
            'Fiction',
            '5',
            '5',
            'Example Publisher',
            '2024-01-01',
            '350',
            'en',
            'example_book.jpg',
            'example_book.pdf'
        ],
        [
            '978-0-987654-32-1',
            'Another Example',
            'Jane Smith',
            'Another sample description',
            'Science',
            '3',
            '3',
            'Science Press',
            '2024-02-15',
            '420',
            'en',
            'another_example.jpg',
            'another_example.pdf'
        ]
    ]
    
    if format_type == 'excel':
        # Generate Excel file
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Book Import Template"
        
        # Write headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write example data
        for row_num, row_data in enumerate(example_data, 2):
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
        
        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="book_import_template.xlsx"'
        workbook.save(response)
        return response
    
    else:
        # Generate CSV file
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="book_import_template.csv"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        
        for row_data in example_data:
            writer.writerow(row_data)
        
        return response
