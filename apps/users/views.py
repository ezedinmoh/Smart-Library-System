from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q
from django.views.decorators.csrf import ensure_csrf_cookie
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import timedelta
from .models import User, UserProfile
from .permissions import admin_required, check_admin
from .notifications import get_user_notifications, get_notification_count
from .forms import CustomUserCreationForm, UserEditForm, UserProfileForm, AdminUserCreationForm
import io
from .models import User, UserProfile
from .forms import CustomUserCreationForm
from .permissions import admin_required, check_admin


def check_username_availability(request):
    """Check if username is available"""
    username = request.GET.get('username', '').strip()
    if not username:
        return JsonResponse({'available': True})
    
    exists = User.objects.filter(username__iexact=username).exists()
    return JsonResponse({'available': not exists})


def check_email_availability(request):
    """Check if email is available"""
    email = request.GET.get('email', '').strip()
    if not email:
        return JsonResponse({'available': True})
    
    exists = User.objects.filter(email__iexact=email).exists()
    return JsonResponse({'available': not exists})


def register(request):
    """User registration view with email verification"""
    if request.user.is_authenticated:
        # Redirect authenticated users to their role-specific dashboard
        if request.user.is_admin:
            return redirect('dashboard:admin')
        elif request.user.is_librarian:
            return redirect('dashboard:librarian')
        elif request.user.is_student:
            return redirect('dashboard:student')

    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        
        if not form.is_valid():
            messages.error(request, 'Please correct the errors below.')
        else:
            user = form.save(commit=False)
            user.is_active = False  # User must verify email first
            user.save()
            
            # Create UserProfile
            UserProfile.objects.get_or_create(user=user)
            
            # Send verification email using django-allauth
            from allauth.account.models import EmailAddress, EmailConfirmation
            from django.core.mail import send_mail
            from django.conf import settings
            
            # Create email address record
            email_address, created = EmailAddress.objects.get_or_create(
                user=user,
                email=user.email,
                defaults={'primary': True, 'verified': False}
            )
            
            # Create confirmation key and build verification URL
            import secrets
            from django.utils import timezone as tz
            from django.urls import reverse

            custom_key = secrets.token_urlsafe(48).replace('=', '').replace('-', '').replace('_', '')[:64]
            confirmation = EmailConfirmation(
                email_address=email_address,
                created=tz.now(),
                key=custom_key
            )
            confirmation.sent = tz.now()
            confirmation.save()

            activate_url = request.build_absolute_uri(
                reverse('users:confirm_email', kwargs={'key': confirmation.key})
            )

            # Try sending email with 10-second timeout
            # If it succeeds → show "check inbox", if it fails/times out → show fallback link
            import socket
            import logging
            from django.core.mail import EmailMultiAlternatives
            from django.conf import settings as django_settings

            subject = f'Verify your email for {settings.SITE_NAME}'
            from django.template.loader import render_to_string
            ctx = {
                'username': user.username,
                'activate_url': activate_url,
                'site_name': settings.SITE_NAME,
                'site_url': settings.SITE_URL,
                'admin_created': False,
            }
            text_content = render_to_string('emails/email_verification.txt', ctx)
            html_content = render_to_string('emails/email_verification.html', ctx)

            email_sent = False
            email_error = None
            old_timeout = socket.getdefaulttimeout()
            try:
                socket.setdefaulttimeout(5)
                msg = EmailMultiAlternatives(subject, text_content, django_settings.DEFAULT_FROM_EMAIL, [user.email])
                msg.attach_alternative(html_content, "text/html")
                msg.send(fail_silently=False)
                email_sent = True
            except Exception as e:
                email_error = str(e)
                logging.getLogger(__name__).warning(f"Verification email failed for {user.email}: {e}")
            finally:
                socket.setdefaulttimeout(old_timeout)

            # Always show verification page - with link visible if email failed
            return render(request, 'account/verification_sent.html', {
                'email': user.email,
                'email_sent': email_sent,
                'email_error': email_error,
                'activate_url': activate_url,  # fallback: show link on page
                'username': user.username,
            })
    else:
        form = CustomUserCreationForm()

    return render(request, 'users/register.html', {'form': form})


@ensure_csrf_cookie
def login_view(request):
    """User login view with proper CSRF handling"""
    if request.user.is_authenticated:
        # Redirect authenticated users to their role-specific dashboard
        if request.user.is_admin:
            return redirect('dashboard:admin')
        elif request.user.is_librarian:
            return redirect('dashboard:librarian')
        elif request.user.is_student:
            return redirect('dashboard:student')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Check if user exists and is inactive (case-insensitive)
        try:
            user_check = User.objects.get(username__iexact=username)
            if not user_check.is_active:
                from django.urls import reverse
                resend_url = reverse('users:resend_verification')
                messages.error(
                    request, 
                    f'Please verify your email address before logging in. '
                    f'<a href="{resend_url}" style="color: #10b981; text-decoration: underline; font-weight: 600;">Resend verification email</a>',
                    extra_tags='safe'
                )
                return render(request, 'users/login.html', {'show_resend_link': True, 'user_email': user_check.email})
        except User.DoesNotExist:
            pass
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome back, {user.username}!')
            
            # Redirect to role-specific dashboard
            if user.is_admin:
                return redirect('dashboard:admin')
            elif user.is_librarian:
                return redirect('dashboard:librarian')
            elif user.is_student:
                return redirect('dashboard:student')
            else:
                return redirect('dashboard:home')
        else:
            messages.error(request, 'Invalid username or password. Please try again.')
            # Return to login page with error message
            return render(request, 'users/login.html')
    
    return render(request, 'users/login.html')


def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('dashboard:home')


@login_required
def profile_view(request):
    user = request.user
    profile, created = UserProfile.objects.get_or_create(user=user)
    
    # Calculate valid until date (4 years from now)
    valid_until = timezone.now() + timedelta(days=365*4)
    
    context = {
        'user': user,
        'profile': profile,
        'valid_until': valid_until,
    }
    
    return render(request, 'users/profile.html', context)


@login_required
def profile_edit(request):
    user = request.user
    profile, created = UserProfile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        user_form = UserEditForm(request.POST, instance=user)
        profile_form = UserProfileForm(request.POST, request.FILES, instance=profile)
        
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, 'Your profile has been updated successfully!')
            return redirect('users:profile')
    else:
        user_form = UserEditForm(instance=user)
        profile_form = UserProfileForm(instance=profile)
    
    context = {
        'user_form': user_form,
        'profile_form': profile_form,
    }
    
    return render(request, 'users/profile_edit.html', context)


@login_required
def change_password(request):
    """Change user password"""
    from django.contrib.auth.forms import PasswordChangeForm
    from django.contrib.auth import update_session_auth_hash
    
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Important: Update the session to prevent logout
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password has been changed successfully!')
            return redirect('users:profile')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(request.user)
    
    # Add Bootstrap classes to form fields
    for field in form.fields.values():
        field.widget.attrs['class'] = 'form-control'
    
    return render(request, 'users/change_password.html', {'form': form})


@login_required
def notifications_view(request):
    """View all notifications for the user"""
    from .notifications import mark_all_notifications_read
    
    # Show non-deleted notifications (deleted ones are permanently removed from view)
    # This shows both read and unread notifications, but not deleted ones
    notifications = get_user_notifications(request.user, include_deleted=False)
    
    # Mark all as read when viewing the notifications page
    if request.GET.get('mark_read') == 'all':
        mark_all_notifications_read(request.user)
        messages.success(request, 'All notifications marked as read.')
        return redirect('users:notifications')
    
    context = {
        'notifications': notifications,
        'notification_count': len(notifications),
        'unread_count': get_notification_count(request.user)
    }
    return render(request, 'users/notifications.html', context)


@login_required
def mark_notification_read_view(request, notification_key):
    """Mark a single notification as read"""
    from .notifications import mark_notification_read
    from django.http import JsonResponse
    
    if request.method == 'POST':
        notification_type = request.POST.get('type', '')
        mark_notification_read(request.user, notification_key, notification_type)
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'count': get_notification_count(request.user)})
        
        return redirect(request.META.get('HTTP_REFERER', 'dashboard:home'))
    
    return redirect('dashboard:home')


@login_required
def delete_notification_view(request, notification_key):
    """Delete a single notification"""
    from .notifications import delete_notification
    from django.http import JsonResponse
    
    if request.method == 'POST':
        notification_type = request.POST.get('type', '')
        success = delete_notification(request.user, notification_key, notification_type)
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': success,
                'unread_count': get_notification_count(request.user)
            })
        
        if success:
            messages.success(request, 'Notification deleted successfully.')
        else:
            messages.error(request, 'Failed to delete notification.')
        
        return redirect(request.META.get('HTTP_REFERER', 'users:notifications'))
    
    return redirect('users:notifications')


@login_required
def clear_all_notifications_view(request):
    """Clear all notifications for the user"""
    from .notifications import clear_all_notifications
    from django.http import JsonResponse
    
    if request.method == 'POST':
        count = clear_all_notifications(request.user)
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'count': count
            })
        
        messages.success(request, f'Successfully deleted {count} notification(s).')
        return redirect('users:notifications')
    
    return redirect('users:notifications')


def check_admin(user):
    return user.is_admin


@login_required
@user_passes_test(check_admin)
def users_list(request):
    """Admin user management with search, filter, pagination"""
    from django.core.paginator import Paginator

    users = User.objects.select_related('profile').order_by('-created_at')

    # Search
    search = request.GET.get('search', '').strip()
    if search:
        users = users.filter(
            Q(username__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search)
        )

    # Role filter
    role_filter = request.GET.get('role', '')
    if role_filter in ['admin', 'librarian', 'student']:
        users = users.filter(role=role_filter)

    # Status filter
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)

    total_count = users.count()
    paginator = Paginator(users, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Summary counts
    from django.db.models import Count
    role_counts = {
        'admin': User.objects.filter(role='admin').count(),
        'librarian': User.objects.filter(role='librarian').count(),
        'student': User.objects.filter(role='student').count(),
        'total': User.objects.count(),
    }

    context = {
        'page_obj': page_obj,
        'users': page_obj.object_list,
        'search': search,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'total_count': total_count,
        'role_counts': role_counts,
    }
    return render(request, 'users/users_list.html', context)


@login_required
@user_passes_test(check_admin)
def user_detail(request, pk):
    """Admin user detail with borrow history and stats"""
    from apps.borrow.models import BorrowRecord, BookRequest
    from django.db.models import Sum

    target_user = get_object_or_404(User, pk=pk)
    profile, _ = UserProfile.objects.get_or_create(user=target_user)

    # Calculate valid until date (4 years from now)
    valid_until = timezone.now() + timedelta(days=365*4)

    # Borrow stats
    borrow_records = BorrowRecord.objects.filter(user=target_user).select_related('book').order_by('-borrow_date')
    active_borrows = borrow_records.filter(status__in=['borrowed', 'overdue'])
    overdue_borrows = borrow_records.filter(status='overdue')
    total_borrowed = borrow_records.count()
    total_returned = borrow_records.filter(status='returned').count()

    # Fine stats
    total_fines = borrow_records.filter(fine_amount__gt=0).aggregate(t=Sum('fine_amount'))['t'] or 0
    unpaid_fines = borrow_records.filter(fine_amount__gt=0, fine_paid=False).aggregate(t=Sum('fine_amount'))['t'] or 0

    # Requests
    pending_requests = BookRequest.objects.filter(user=target_user, status='pending').select_related('book')

    context = {
        'target_user': target_user,
        'profile': profile,
        'valid_until': valid_until,
        'borrow_records': borrow_records[:10],
        'active_borrows': active_borrows,
        'overdue_borrows': overdue_borrows,
        'pending_requests': pending_requests,
        'total_borrowed': total_borrowed,
        'total_returned': total_returned,
        'total_fines': total_fines,
        'unpaid_fines': unpaid_fines,
    }
    return render(request, 'users/user_detail.html', context)


@login_required
@user_passes_test(check_admin)
def user_role_change(request, pk):
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        new_role = request.POST.get('role')
        if new_role in dict(User.ROLE_CHOICES):
            user.role = new_role
            user.save()
            messages.success(request, f'{user.username} role changed to {user.get_role_display()}.')
            return redirect('users:detail', pk=user.pk)
    
    context = {
        'user': user,
        'roles': User.ROLE_CHOICES,
    }
    
    return render(request, 'users/user_role_change.html', context)


@login_required
@user_passes_test(check_admin)
def update_borrow_limit(request, pk):
    """Update user's borrow limit"""
    user = get_object_or_404(User, pk=pk)
    profile, _ = UserProfile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        max_books = request.POST.get('max_books_allowed')
        try:
            max_books = int(max_books)
            if 1 <= max_books <= 20:
                profile.max_books_allowed = max_books
                profile.save()
                messages.success(request, f'Borrow limit updated to {max_books} books for {user.username}.')
            else:
                messages.error(request, 'Borrow limit must be between 1 and 20.')
        except (ValueError, TypeError):
            messages.error(request, 'Invalid borrow limit value.')
    
    return redirect('users:detail', pk=user.pk)


@login_required
@user_passes_test(check_admin)
def user_create(request):
    """Admin creates a new user with email verification"""
    if request.method == 'POST':
        form = AdminUserCreationForm(request.POST)  # Use AdminUserCreationForm
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False  # User must verify email first
            user.save()
            
            # Create UserProfile
            UserProfile.objects.get_or_create(user=user)
            
            # Send verification email
            from allauth.account.models import EmailAddress, EmailConfirmation
            from django.core.mail import EmailMultiAlternatives
            from django.conf import settings
            import secrets
            
            # Create email address record
            email_address, created = EmailAddress.objects.get_or_create(
                user=user,
                email=user.email,
                defaults={'primary': True, 'verified': False}
            )
            
            # Create confirmation with custom key
            custom_key = secrets.token_urlsafe(48).replace('=', '').replace('-', '').replace('_', '')[:64]
            
            from django.utils import timezone as tz
            confirmation = EmailConfirmation(
                email_address=email_address,
                created=tz.now(),
                key=custom_key
            )
            confirmation.sent = tz.now()
            confirmation.save()
            
            # Build activation URL
            from django.urls import reverse
            activate_url = request.build_absolute_uri(
                reverse('users:confirm_email', kwargs={'key': confirmation.key})
            )
            
            # Email content
            subject = f'Verify your email for {settings.SITE_NAME}'
            from django.template.loader import render_to_string
            ctx = {
                'username': user.username,
                'activate_url': activate_url,
                'site_name': settings.SITE_NAME,
                'site_url': settings.SITE_URL,
                'admin_created': True,
            }
            plain_message = render_to_string('emails/email_verification.txt', ctx)
            html_message  = render_to_string('emails/email_verification.html', ctx)
            
            # Send email
            try:
                email = EmailMultiAlternatives(
                    subject=subject,
                    body=plain_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[user.email]
                )
                email.attach_alternative(html_message, "text/html")
                email.send(fail_silently=False)
                
                # Show verification sent page instead of redirecting
                return render(request, 'users/user_created_verification_sent.html', {
                    'created_user': user,
                    'verification_email': user.email
                })
            except Exception as e:
                messages.warning(request, f'User "{user.username}" created, but failed to send verification email: {str(e)}')
                return redirect('users:detail', pk=user.pk)
    else:
        form = AdminUserCreationForm()  # Use AdminUserCreationForm for role selection
    return render(request, 'users/user_create.html', {'form': form})


@login_required
@admin_required
def export_users_csv(request):
    """Export users to CSV (Admin only)"""
    from django.http import HttpResponse
    import csv
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="users_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Username', 'Email', 'First Name', 'Last Name', 'Role', 
        'Phone Number', 'Address', 'Is Active', 'Date Joined',
        'Currently Borrowed', 'Max Books Allowed', 'Total Fines'
    ])
    
    users = User.objects.select_related('profile').all()
    
    for user in users:
        writer.writerow([
            user.username,
            user.email,
            user.first_name,
            user.last_name,
            user.get_role_display(),
            user.phone_number,
            user.address,
            'Yes' if user.is_active else 'No',
            user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
            user.profile.currently_borrowed if hasattr(user, 'profile') else 0,
            user.profile.max_books_allowed if hasattr(user, 'profile') else 0,
            user.profile.total_fines if hasattr(user, 'profile') else 0
        ])
    
    return response

@login_required
@admin_required
def export_users_excel(request):
    """Export users to Excel (Admin only)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    # Define headers
    headers = [
        'Username', 'Email', 'First Name', 'Last Name', 'Role', 
        'Phone Number', 'Address', 'Is Active', 'Date Joined',
        'Currently Borrowed', 'Max Books Allowed', 'Total Fines'
    ]
    
    # Add headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add data
    users = User.objects.select_related('profile').all()
    
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.username)
        ws.cell(row=row, column=2, value=user.email)
        ws.cell(row=row, column=3, value=user.first_name)
        ws.cell(row=row, column=4, value=user.last_name)
        ws.cell(row=row, column=5, value=user.get_role_display())
        ws.cell(row=row, column=6, value=user.phone_number)
        ws.cell(row=row, column=7, value=user.address)
        ws.cell(row=row, column=8, value='Yes' if user.is_active else 'No')
        ws.cell(row=row, column=9, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=10, value=user.profile.currently_borrowed if hasattr(user, 'profile') else 0)
        ws.cell(row=row, column=11, value=user.profile.max_books_allowed if hasattr(user, 'profile') else 0)
        ws.cell(row=row, column=12, value=float(user.profile.total_fines) if hasattr(user, 'profile') else 0)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="users_export.xlsx"'
    
    wb.save(response)
    return response


@login_required
def upload_avatar(request):
    """Upload profile picture via AJAX"""
    from django.http import JsonResponse
    import os
    from django.core.files.storage import default_storage
    
    if request.method == 'POST' and request.FILES.get('profile_picture'):
        try:
            profile, created = UserProfile.objects.get_or_create(user=request.user)
            
            # Delete old profile picture if exists (local storage only)
            if profile.profile_picture:
                try:
                    storage = profile.profile_picture.storage
                    is_cloudinary = 'cloudinary' in type(storage).__module__.lower()
                    if not is_cloudinary:
                        import os
                        if os.path.isfile(profile.profile_picture.path):
                            os.remove(profile.profile_picture.path)
                except Exception:
                    pass
            
            # Save new profile picture
            profile.profile_picture = request.FILES['profile_picture']
            profile.save()
            
            return JsonResponse({
                'success': True,
                'avatar_url': profile.profile_picture.url,
                'message': 'Profile picture updated successfully'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request'
    }, status=400)


@login_required
def update_profile(request):
    """Update user profile information via AJAX"""
    from django.http import JsonResponse
    
    if request.method == 'POST':
        try:
            user = request.user
            
            # Update user fields
            user.first_name = request.POST.get('first_name', '').strip()
            user.last_name = request.POST.get('last_name', '').strip()
            user.email = request.POST.get('email', '').strip()
            user.phone_number = request.POST.get('phone_number', '').strip()
            user.address = request.POST.get('address', '').strip()
            
            # Validate email
            if not user.email:
                return JsonResponse({
                    'success': False,
                    'error': 'Email is required'
                }, status=400)
            
            # Check if email is already taken by another user
            if User.objects.filter(email=user.email).exclude(id=user.id).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Email is already taken'
                }, status=400)
            
            user.save()
            
            return JsonResponse({
                'success': True,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'phone_number': user.phone_number,
                'address': user.address,
                'message': 'Profile updated successfully'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request'
    }, status=400)


@login_required
def change_password_ajax(request):
    """Change user password via AJAX"""
    from django.http import JsonResponse
    from django.contrib.auth import update_session_auth_hash
    
    if request.method == 'POST':
        try:
            user = request.user
            current_password = request.POST.get('current_password', '')
            new_password = request.POST.get('new_password', '')
            
            # Verify current password
            if not user.check_password(current_password):
                return JsonResponse({
                    'success': False,
                    'error': 'Current password is incorrect'
                }, status=400)
            
            # Validate new password
            if len(new_password) < 8:
                return JsonResponse({
                    'success': False,
                    'error': 'Password must be at least 8 characters long'
                }, status=400)
            
            # Set new password
            user.set_password(new_password)
            user.save()
            
            # Update session to prevent logout
            update_session_auth_hash(request, user)
            
            # Send password changed notification email
            try:
                from django.core.mail import EmailMultiAlternatives
                from django.conf import settings
                
                subject = f'Password Changed - {settings.SITE_NAME}'
                plain_message = f"""
================================================================================
PASSWORD CHANGED - {settings.SITE_NAME}
================================================================================

Hello {user.get_full_name() or user.username},

Your password was successfully changed on {timezone.now().strftime('%B %d, %Y at %I:%M %p')}.

If you did not make this change, please reset your password immediately and contact support.

Best regards,
{settings.SITE_NAME} Team

================================================================================
"""
                
                html_message = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px;">
    <div style="background: linear-gradient(135deg, #10b981, #059669); padding: 30px; border-radius: 10px 10px 0 0; text-align: center;">
        <h1 style="color: white; margin: 0; font-size: 24px;">🔐 Password Changed</h1>
    </div>
    
    <div style="background: #f9fafb; padding: 30px; border-radius: 0 0 10px 10px;">
        <h2 style="color: #1f2937; margin-top: 0;">Hello {user.get_full_name() or user.username},</h2>
        
        <p style="color: #4b5563; font-size: 16px;">Your password was successfully changed on <strong>{timezone.now().strftime('%B %d, %Y at %I:%M %p')}</strong>.</p>
        
        <div style="background: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin: 20px 0; border-radius: 4px;">
            <p style="color: #92400e; margin: 0; font-size: 14px;"><strong>⚠️ Security Notice:</strong> If you did not make this change, please reset your password immediately and contact support.</p>
        </div>
        
        <hr style="margin: 30px 0; border: none; border-top: 1px solid #d1d5db;">
        
        <p style="color: #9ca3af; font-size: 12px; text-align: center;">Best regards,<br>{settings.SITE_NAME} Team</p>
    </div>
</body>
</html>
"""
                
                email = EmailMultiAlternatives(
                    subject=subject,
                    body=plain_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[user.email]
                )
                email.attach_alternative(html_message, "text/html")
                email.send(fail_silently=True)
            except Exception as e:
                # Log error but don't fail the password change
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to send password change email to {user.email}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'message': 'Password changed successfully'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request'
    }, status=400)


# ============================================================================
# BULK OPERATIONS
# ============================================================================

@login_required
@admin_required
def bulk_import_users(request):
    """Bulk import users from CSV/Excel file"""
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        
        if not uploaded_file:
            messages.error(request, 'Please select a file to upload.')
            return redirect('users:bulk_import')
        
        # Check file extension
        file_extension = uploaded_file.name.split('.')[-1].lower()
        
        if file_extension not in ['csv', 'xlsx', 'xls']:
            messages.error(request, 'Invalid file format. Please upload CSV or Excel file.')
            return redirect('users:bulk_import')
        
        try:
            imported_count = 0
            skipped_count = 0
            errors = []
            
            if file_extension == 'csv':
                # Process CSV file
                import csv
                import io
                
                decoded_file = uploaded_file.read().decode('utf-8')
                csv_reader = csv.DictReader(io.StringIO(decoded_file))
                
                for row_num, row in enumerate(csv_reader, start=2):
                    try:
                        username = row.get('username', '').strip()
                        email = row.get('email', '').strip()
                        
                        # Check if user already exists
                        if User.objects.filter(username=username).exists():
                            skipped_count += 1
                            errors.append(f"Row {row_num}: Username '{username}' already exists")
                            continue
                        
                        if User.objects.filter(email=email).exists():
                            skipped_count += 1
                            errors.append(f"Row {row_num}: Email '{email}' already exists")
                            continue
                        
                        # Create user
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            password=row.get('password', 'changeme123'),
                            first_name=row.get('first_name', '').strip(),
                            last_name=row.get('last_name', '').strip(),
                            phone_number=row.get('phone_number', '').strip(),
                            address=row.get('address', '').strip(),
                            role=row.get('role', 'student').strip().lower(),
                        )
                        
                        # Create profile
                        UserProfile.objects.get_or_create(user=user)
                        
                        imported_count += 1
                        
                    except Exception as e:
                        skipped_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")
            
            else:
                # Process Excel file
                import openpyxl
                
                workbook = openpyxl.load_workbook(uploaded_file)
                sheet = workbook.active
                
                # Get headers from first row
                headers = [cell.value for cell in sheet[1]]
                
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Create dictionary from headers and row values
                        row_dict = dict(zip(headers, row))
                        
                        username = str(row_dict.get('username', '')).strip()
                        email = str(row_dict.get('email', '')).strip()
                        
                        # Check if user already exists
                        if User.objects.filter(username=username).exists():
                            skipped_count += 1
                            errors.append(f"Row {row_num}: Username '{username}' already exists")
                            continue
                        
                        if User.objects.filter(email=email).exists():
                            skipped_count += 1
                            errors.append(f"Row {row_num}: Email '{email}' already exists")
                            continue
                        
                        # Create user
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            password=str(row_dict.get('password', 'changeme123')),
                            first_name=str(row_dict.get('first_name', '')).strip(),
                            last_name=str(row_dict.get('last_name', '')).strip(),
                            phone_number=str(row_dict.get('phone_number', '')).strip(),
                            address=str(row_dict.get('address', '')).strip(),
                            role=str(row_dict.get('role', 'student')).strip().lower(),
                        )
                        
                        # Create profile
                        UserProfile.objects.get_or_create(user=user)
                        
                        imported_count += 1
                        
                    except Exception as e:
                        skipped_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")
            
            # Log activity
            from apps.dashboard.utils import log_activity
            log_activity(
                request.user, 
                'user_created', 
                f'Bulk imported {imported_count} users from {uploaded_file.name}', 
                request
            )
            
            # Show results
            if imported_count > 0:
                messages.success(request, f'Successfully imported {imported_count} user(s).')
            if skipped_count > 0:
                messages.warning(request, f'Skipped {skipped_count} user(s). Check errors below.')
            
            # Store errors in session for display
            if errors:
                request.session['bulk_import_errors'] = errors[:50]  # Limit to 50 errors
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
        
        return redirect('users:bulk_import')
    
    # GET request - show upload form
    errors = request.session.pop('bulk_import_errors', [])
    
    context = {
        'errors': errors,
    }
    return render(request, 'users/bulk_import.html', context)


@login_required
@login_required
@admin_required
def download_user_import_template(request):
    """Download CSV or Excel template for bulk user import"""
    format_type = request.GET.get('format', 'csv').lower()
    
    # Template data
    headers = [
        'username', 'email', 'password', 'first_name', 'last_name', 
        'phone_number', 'address', 'role'
    ]
    
    example_data = [
        [
            'john_doe',
            'john.doe@example.com',
            'changeme123',
            'John',
            'Doe',
            '+1234567890',
            '123 Main St, City',
            'student'
        ],
        [
            'jane_smith',
            'jane.smith@example.com',
            'changeme123',
            'Jane',
            'Smith',
            '+0987654321',
            '456 Oak Ave, Town',
            'student'
        ],
        [
            'librarian_user',
            'librarian@example.com',
            'changeme123',
            'Library',
            'Staff',
            '+1122334455',
            '789 Library Rd',
            'librarian'
        ]
    ]
    
    if format_type == 'excel':
        # Generate Excel file
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "User Import Template"
        
        # Write headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0ea5e9", end_color="0ea5e9", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write example data
        for row_num, row_data in enumerate(example_data, 2):
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
        
        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="user_import_template.xlsx"'
        workbook.save(response)
        return response
    
    else:
        # Generate CSV file
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="user_import_template.csv"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        
        for row_data in example_data:
            writer.writerow(row_data)
        
        return response


@login_required
@admin_required
def bulk_email_users(request):
    """Send bulk email to selected users"""
    if request.method == 'POST':
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        recipient_type = request.POST.get('recipient_type', 'all')
        
        if not subject or not message:
            messages.error(request, 'Subject and message are required.')
            return redirect('users:bulk_email')
        
        # Get recipients based on type
        if recipient_type == 'all':
            recipients = User.objects.filter(is_active=True)
        elif recipient_type == 'students':
            recipients = User.objects.filter(role='student', is_active=True)
        elif recipient_type == 'librarians':
            recipients = User.objects.filter(role='librarian', is_active=True)
        elif recipient_type == 'admins':
            recipients = User.objects.filter(role='admin', is_active=True)
        else:
            messages.error(request, 'Invalid recipient type.')
            return redirect('users:bulk_email')
        
        # Send emails
        from django.core.mail import send_mass_mail
        from django.conf import settings
        
        email_messages = []
        sent_count = 0
        
        for user in recipients:
            if user.email:
                # Personalize message
                personalized_message = message.replace('{name}', user.get_full_name() or user.username)
                personalized_message = personalized_message.replace('{username}', user.username)
                
                email_messages.append((
                    subject,
                    personalized_message,
                    settings.DEFAULT_FROM_EMAIL,
                    [user.email]
                ))
                sent_count += 1
        
        try:
            send_mass_mail(email_messages, fail_silently=False)
            
            # Log activity
            from apps.dashboard.utils import log_activity
            log_activity(
                request.user, 
                'other', 
                f'Sent bulk email to {sent_count} users: "{subject}"', 
                request
            )
            
            messages.success(request, f'Successfully sent email to {sent_count} user(s).')
        except Exception as e:
            messages.error(request, f'Error sending emails: {str(e)}')
        
        return redirect('users:bulk_email')
    
    # GET request - show form
    context = {
        'total_users': User.objects.filter(is_active=True).count(),
        'total_students': User.objects.filter(role='student', is_active=True).count(),
        'total_librarians': User.objects.filter(role='librarian', is_active=True).count(),
        'total_admins': User.objects.filter(role='admin', is_active=True).count(),
    }
    return render(request, 'users/bulk_email.html', context)


def confirm_email(request, key):
    """Custom email confirmation view"""
    from allauth.account.models import EmailConfirmation, EmailAddress
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Remove any URL encoding artifacts
    import urllib.parse
    key = urllib.parse.unquote(key)
    
    # Debug logging
    logger.info(f"Email verification attempt with key: {key}")
    logger.info(f"Key length: {len(key)}")
    
    try:
        # Get the confirmation object (try exact match first, then variations)
        confirmation = None
        try:
            confirmation = EmailConfirmation.objects.get(key=key)
            logger.info(f"Found confirmation with exact key match")
        except EmailConfirmation.DoesNotExist:
            # Try without trailing =
            try:
                confirmation = EmailConfirmation.objects.get(key=key.rstrip('='))
                logger.info(f"Found confirmation after stripping trailing =")
            except EmailConfirmation.DoesNotExist:
                # Try lowercase
                try:
                    confirmation = EmailConfirmation.objects.get(key=key.lower())
                    logger.info(f"Found confirmation with lowercase key")
                except EmailConfirmation.DoesNotExist:
                    try:
                        confirmation = EmailConfirmation.objects.get(key=key.lower().rstrip('='))
                        logger.info(f"Found confirmation with lowercase key and stripped =")
                    except EmailConfirmation.DoesNotExist:
                        # Log all existing keys for debugging
                        all_keys = EmailConfirmation.objects.values_list('key', flat=True)
                        logger.error(f"No match found. Existing keys in database: {list(all_keys)}")
                        raise
        
        # Check if already confirmed
        if confirmation.email_address.verified:
            messages.info(request, 'This email has already been verified.')
            return redirect('users:login')
        
        # Confirm email immediately on GET (one-click verification)
        from django.db import transaction

        email_address = confirmation.email_address
        user = email_address.user

        try:
            with transaction.atomic():
                EmailAddress.objects.filter(
                    email=email_address.email,
                    verified=True
                ).exclude(pk=email_address.pk).delete()

                email_address.verified = True
                email_address.primary = True
                email_address.save()

                user.is_active = True
                user.save()

        except Exception as e:
            logger.error(f"Error during email confirmation for {user.username}: {e}")
            user.is_active = True
            user.save()
            EmailAddress.objects.filter(pk=email_address.pk).update(verified=True, primary=True)

        # Send welcome email
        try:
            from .notifications import send_welcome_email
            send_welcome_email(user)
        except Exception:
            pass

        # Delete the confirmation
        try:
            confirmation.delete()
        except Exception:
            pass

        logger.info(f"Successfully verified email for user: {user.username}")

        return render(request, 'account/email_confirmed.html', {
            'verified_user': user
        })
    
    except EmailConfirmation.DoesNotExist:
        logger.error(f"EmailConfirmation not found for key: {key}")
        messages.error(request, 'Invalid or expired verification link.')
        return redirect('users:register')



def resend_verification_email(request):
    """Resend verification email for unverified users"""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        
        if not email:
            messages.error(request, 'Please enter your email address.')
            return render(request, 'users/resend_verification.html')
        
        try:
            # Find user by email
            user = User.objects.get(email=email)
            
            # Check if already verified
            if user.is_active:
                messages.info(request, 'This email is already verified. You can login now.')
                return redirect('users:login')
            
            # Check if email address record exists
            from allauth.account.models import EmailAddress, EmailConfirmation
            email_address = EmailAddress.objects.filter(user=user, email=email).first()
            
            if not email_address:
                # Create email address record
                email_address = EmailAddress.objects.create(
                    user=user,
                    email=email,
                    primary=True,
                    verified=False
                )
            
            # Delete old confirmations for this email
            EmailConfirmation.objects.filter(email_address=email_address).delete()
            
            # Create new confirmation with custom key
            import secrets
            custom_key = secrets.token_urlsafe(48).replace('=', '').replace('-', '').replace('_', '')[:64]
            
            from django.utils import timezone as tz
            confirmation = EmailConfirmation(
                email_address=email_address,
                created=tz.now(),
                key=custom_key
            )
            confirmation.sent = tz.now()
            confirmation.save()
            
            # Build activation URL
            from django.urls import reverse
            activate_url = request.build_absolute_uri(
                reverse('users:confirm_email', kwargs={'key': confirmation.key})
            )
            
            # Send email
            from django.core.mail import EmailMultiAlternatives
            from django.conf import settings
            from django.template.loader import render_to_string

            subject = f'Verify your email for {settings.SITE_NAME}'
            ctx = {
                'username': user.username,
                'activate_url': activate_url,
                'site_name': settings.SITE_NAME,
                'site_url': settings.SITE_URL,
                'admin_created': False,
            }
            plain_message = render_to_string('emails/email_verification.txt', ctx)
            html_message  = render_to_string('emails/email_verification.html', ctx)

            email_msg = EmailMultiAlternatives(
                subject=subject,
                body=plain_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[user.email]
            )
            email_msg.attach_alternative(html_message, "text/html")
            email_msg.send(fail_silently=False)
            
            messages.success(request, f'Verification email has been resent to {email}. Please check your inbox.')
            return redirect('users:login')
            
        except User.DoesNotExist:
            # Don't reveal if email exists (security)
            messages.success(request, f'If an account exists with this email, a verification link has been sent.')
            return redirect('users:login')
        except Exception as e:
            messages.error(request, f'Failed to send verification email: {str(e)}')
            return render(request, 'users/resend_verification.html')
    
    return render(request, 'users/resend_verification.html')
